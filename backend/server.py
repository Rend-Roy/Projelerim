from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Depends, Header
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
from fpdf import FPDF
import io
from openpyxl import load_workbook
from passlib.context import CryptContext
import jwt

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# =============================================================================
# FAZ 3.0: Authentication Configuration
# =============================================================================
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer(auto_error=False)

# JWT Configuration
JWT_SECRET = os.environ.get("JWT_SECRET", "satis-takip-secret-key-2024")
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24 * 7  # 7 gün (Beni Hatırla için)

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(user_id: str, email: str) -> str:
    expire = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    payload = {
        "sub": user_id,
        "email": email,
        "exp": expire
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token süresi dolmuş")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Geçersiz token")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> Optional[dict]:
    """Mevcut kullanıcıyı al - opsiyonel (geriye dönük uyumluluk için)"""
    if not credentials:
        return None
    try:
        payload = decode_token(credentials.credentials)
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        return user
    except:
        return None

async def require_auth(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    """Zorunlu authentication - giriş yapılmamışsa hata döndür"""
    if not credentials:
        raise HTTPException(status_code=401, detail="Giriş yapmanız gerekiyor")
    try:
        payload = decode_token(credentials.credentials)
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Kullanıcı bulunamadı")
        return user
    except HTTPException:
        raise
    except:
        raise HTTPException(status_code=401, detail="Geçersiz oturum")

# =============================================================================
# FAZ 3.0: User Model
# =============================================================================
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    password_hash: str
    name: str
    role: str = "representative"  # representative, admin (ileride)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserRegister(BaseModel):
    email: str
    password: str
    name: str

class UserLogin(BaseModel):
    email: str
    password: str
    remember_me: bool = False

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    created_at: datetime

class ForgotPasswordRequest(BaseModel):
    email: str

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str

# =============================================================================
# Define Models (Mevcut modeller - user_id eklendi)
# =============================================================================
class Customer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    region: str
    phone: Optional[str] = None
    address: Optional[str] = None
    price_status: str = "Standart"  # İskontolu veya Standart
    visit_days: List[str] = []  # Pazartesi, Salı, Çarşamba, Perşembe, Cuma, Cumartesi, Pazar
    alerts: List[str] = []  # Müşteri uyarıları (Kırmızı Bayrak)
    user_id: Optional[str] = None  # FAZ 3.0: Kullanıcıya ait (opsiyonel - geriye uyumluluk)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Müşteri uyarı seçenekleri
CUSTOMER_ALERTS = [
    "Geç öder",
    "Fiyat hassas",
    "Belirli saatlerde",
    "Özel anlaşma var",
    "Tahsilat problemi var",
    "Sürekli erteleme yapıyor"
]

class CustomerCreate(BaseModel):
    name: str
    region: str
    phone: Optional[str] = None
    address: Optional[str] = None
    price_status: str = "Standart"
    visit_days: List[str] = []
    alerts: List[str] = []

class CustomerUpdate(BaseModel):
    name: Optional[str] = None
    region: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    price_status: Optional[str] = None
    visit_days: Optional[List[str]] = None
    alerts: Optional[List[str]] = None

# Region models
class Region(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    user_id: Optional[str] = None  # FAZ 3.0: Kullanıcıya ait (opsiyonel)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RegionCreate(BaseModel):
    name: str
    description: Optional[str] = None

class RegionUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None

# Ziyaret durumu seçenekleri
VISIT_STATUS_OPTIONS = ["pending", "visited", "not_visited"]

class Visit(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str
    date: str  # YYYY-MM-DD format
    # Yeni status alanı: pending (bekliyor), visited (ziyaret edildi), not_visited (ziyaret edilmedi)
    status: str = "pending"
    completed: bool = False  # Geriye uyumluluk için korunuyor
    visit_skip_reason: Optional[str] = None  # Ziyaret edilmediyse sebebi
    payment_collected: bool = False  # Tahsilat yapıldı mı
    payment_skip_reason: Optional[str] = None  # Tahsilat yapılmadıysa sebebi
    payment_type: Optional[str] = None  # Nakit, Kredi Kartı, Havale, Çek
    payment_amount: Optional[float] = None  # Tahsilat tutarı
    customer_request: Optional[str] = None  # Müşteri özel talep/notu
    note: Optional[str] = None  # Genel not
    completed_at: Optional[datetime] = None
    # FAZ 2: Ziyaret Süresi ve Kalite
    started_at: Optional[datetime] = None  # Ziyaret başlangıç zamanı
    ended_at: Optional[datetime] = None  # Ziyaret bitiş zamanı
    duration_minutes: Optional[int] = None  # Ziyaret süresi (dakika)
    quality_rating: Optional[int] = None  # Ziyaret kalitesi (1-5 yıldız)
    user_id: Optional[str] = None  # FAZ 3.0: Kullanıcıya ait (opsiyonel)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class VisitUpdate(BaseModel):
    status: Optional[str] = None  # pending, visited, not_visited
    completed: Optional[bool] = None  # Geriye uyumluluk
    visit_skip_reason: Optional[str] = None
    payment_collected: Optional[bool] = None
    payment_skip_reason: Optional[str] = None
    payment_type: Optional[str] = None
    payment_amount: Optional[float] = None
    customer_request: Optional[str] = None
    note: Optional[str] = None
    quality_rating: Optional[int] = None  # 1-5 yıldız

# Follow-Up model
class FollowUp(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str
    due_date: str  # YYYY-MM-DD
    due_time: Optional[str] = None  # HH:MM
    status: str = "pending"  # pending, done, late
    reason: Optional[str] = None
    note: Optional[str] = None
    completed_at: Optional[datetime] = None
    user_id: Optional[str] = None  # FAZ 3.0: Kullanıcıya ait (opsiyonel)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FollowUpCreate(BaseModel):
    customer_id: str
    due_date: str
    due_time: Optional[str] = None
    reason: Optional[str] = None
    note: Optional[str] = None

class FollowUpUpdate(BaseModel):
    due_date: Optional[str] = None
    due_time: Optional[str] = None
    status: Optional[str] = None
    reason: Optional[str] = None
    note: Optional[str] = None

# Daily Report Note model
class DailyReportNote(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str  # YYYY-MM-DD format
    note: str  # Gün sonu genel notu
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DailyReportNoteUpdate(BaseModel):
    note: str

# =============================================================================
# FAZ 4: Araç, Yakıt ve Günlük KM Takibi Modelleri
# =============================================================================

# Yakıt türleri
FUEL_TYPES = ["Benzin", "Dizel", "LPG", "Elektrik", "Hibrit"]

class Vehicle(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    name: str  # Araç adı (ör: Fiat Doblo)
    plate: Optional[str] = None  # Plaka (opsiyonel)
    fuel_type: str = "Benzin"  # Benzin, Dizel, LPG, Elektrik, Hibrit
    starting_km: float = 0  # Başlangıç kilometresi
    is_active: bool = True  # Aktif/pasif durumu
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class VehicleCreate(BaseModel):
    name: str
    plate: Optional[str] = None
    fuel_type: str = "Benzin"
    starting_km: float = 0
    is_active: bool = True

class VehicleUpdate(BaseModel):
    name: Optional[str] = None
    plate: Optional[str] = None
    fuel_type: Optional[str] = None
    starting_km: Optional[float] = None
    is_active: Optional[bool] = None

class FuelRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    vehicle_id: str
    date: str  # YYYY-MM-DD
    current_km: float  # O anki kilometre
    liters: float  # Alınan yakıt (litre veya kWh)
    amount: float  # Toplam tutar (TL)
    note: Optional[str] = None
    # Hesaplanan alanlar
    distance_since_last: Optional[float] = None  # Bir önceki yakıta göre gidilen km
    consumption_per_100km: Optional[float] = None  # 100 km'de tüketim
    cost_per_km: Optional[float] = None  # KM başı maliyet (TL/km)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FuelRecordCreate(BaseModel):
    vehicle_id: str
    date: str
    current_km: float
    liters: float
    amount: float
    note: Optional[str] = None

class DailyKmRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    vehicle_id: str
    date: str  # YYYY-MM-DD
    start_km: float  # Gün başlangıç kilometresi
    end_km: Optional[float] = None  # Gün bitiş kilometresi
    daily_km: Optional[float] = None  # Günlük yapılan km (otomatik hesaplanır)
    avg_cost_per_km: Optional[float] = None  # Ortalama km maliyeti (son 30 gün)
    daily_cost: Optional[float] = None  # Günlük araç maliyeti (TL)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DailyKmRecordCreate(BaseModel):
    vehicle_id: str
    date: str
    start_km: float
    end_km: Optional[float] = None

class DailyKmRecordUpdate(BaseModel):
    start_km: Optional[float] = None
    end_km: Optional[float] = None

# =============================================================================
# FAZ 5: Ürün Kataloğu Modelleri
# =============================================================================

# Birim seçenekleri
PRODUCT_UNITS = ["Adet", "Kg", "Lt", "Paket", "Kutu", "Koli", "Metre", "M²", "M³"]

class Category(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    name: str
    description: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CategoryCreate(BaseModel):
    name: str
    description: Optional[str] = None
    is_active: bool = True

class CategoryUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    is_active: Optional[bool] = None

class Product(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    product_code: str  # UNIQUE per user
    name: str
    category: str  # Kategori adı
    description: Optional[str] = None
    base_price: float = 0  # Sabit fiyat
    unit: str = "Adet"  # Birim
    images: List[str] = []  # Cloudinary URL'leri
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProductCreate(BaseModel):
    product_code: str
    name: str
    category: str
    description: Optional[str] = None
    base_price: float = 0
    unit: str = "Adet"
    images: List[str] = []
    is_active: bool = True

class ProductUpdate(BaseModel):
    product_code: Optional[str] = None
    name: Optional[str] = None
    category: Optional[str] = None
    description: Optional[str] = None
    base_price: Optional[float] = None
    unit: Optional[str] = None
    images: Optional[List[str]] = None
    is_active: Optional[bool] = None

# =============================================================================
# FAZ 3.0: Authentication Endpoints
# =============================================================================

@api_router.post("/auth/register")
async def register(input: UserRegister):
    """Yeni kullanıcı kaydı"""
    # Email kontrolü
    existing = await db.users.find_one({"email": input.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Bu e-posta adresi zaten kayıtlı")
    
    # Şifre validasyonu
    if len(input.password) < 6:
        raise HTTPException(status_code=400, detail="Şifre en az 6 karakter olmalı")
    
    # İlk kullanıcı mı kontrol et (mevcut verileri atamak için)
    user_count = await db.users.count_documents({})
    is_first_user = user_count == 0
    
    # Kullanıcı oluştur
    user = User(
        email=input.email.lower(),
        password_hash=hash_password(input.password),
        name=input.name,
        role="representative"
    )
    
    await db.users.insert_one(user.model_dump())
    
    # İlk kullanıcıysa, mevcut tüm verileri bu kullanıcıya ata
    if is_first_user:
        await db.customers.update_many(
            {"user_id": None},
            {"$set": {"user_id": user.id}}
        )
        await db.visits.update_many(
            {"user_id": None},
            {"$set": {"user_id": user.id}}
        )
        await db.follow_ups.update_many(
            {"user_id": None},
            {"$set": {"user_id": user.id}}
        )
        await db.regions.update_many(
            {"user_id": None},
            {"$set": {"user_id": user.id}}
        )
        logging.info(f"İlk kullanıcı kaydı: Mevcut veriler {user.id} kullanıcısına atandı")
    
    # Token oluştur
    token = create_access_token(user.id, user.email)
    
    return {
        "message": "Kayıt başarılı",
        "token": token,
        "user": {
            "id": user.id,
            "email": user.email,
            "name": user.name,
            "role": user.role
        }
    }

@api_router.post("/auth/login")
async def login(input: UserLogin):
    """Kullanıcı girişi"""
    user = await db.users.find_one({"email": input.email.lower()}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="E-posta veya şifre hatalı")
    
    if not verify_password(input.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="E-posta veya şifre hatalı")
    
    # Token oluştur
    token = create_access_token(user["id"], user["email"])
    
    return {
        "message": "Giriş başarılı",
        "token": token,
        "user": {
            "id": user["id"],
            "email": user["email"],
            "name": user["name"],
            "role": user["role"]
        }
    }

@api_router.post("/auth/logout")
async def logout(current_user: dict = Depends(require_auth)):
    """Kullanıcı çıkışı (client-side token silme)"""
    return {"message": "Çıkış başarılı"}

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(require_auth)):
    """Mevcut kullanıcı bilgisini al"""
    return {
        "id": current_user["id"],
        "email": current_user["email"],
        "name": current_user["name"],
        "role": current_user["role"]
    }

@api_router.post("/auth/forgot-password")
async def forgot_password(input: ForgotPasswordRequest):
    """Şifre sıfırlama talebi (MOCK - konsola yazdırır)"""
    user = await db.users.find_one({"email": input.email.lower()}, {"_id": 0})
    
    # Güvenlik: Kullanıcı olsun olmasın aynı mesajı döndür
    if user:
        # Reset token oluştur (gerçek uygulamada e-posta ile gönderilir)
        reset_token = str(uuid.uuid4())
        expire = datetime.now(timezone.utc) + timedelta(hours=1)
        
        await db.password_resets.insert_one({
            "user_id": user["id"],
            "token": reset_token,
            "expires_at": expire.isoformat(),
            "used": False
        })
        
        # MOCK: Konsola yazdır (gerçek uygulamada e-posta gönderilir)
        logging.info(f"=== ŞİFRE SIFIRLAMA MOCK ===")
        logging.info(f"E-posta: {input.email}")
        logging.info(f"Sıfırlama Token: {reset_token}")
        logging.info(f"Geçerlilik: 1 saat")
        logging.info(f"============================")
        print(f"\n🔐 ŞİFRE SIFIRLAMA TOKEN (MOCK)")
        print(f"   E-posta: {input.email}")
        print(f"   Token: {reset_token}\n")
    
    return {"message": "Şifre sıfırlama bağlantısı e-posta adresinize gönderildi"}

@api_router.post("/auth/reset-password")
async def reset_password(input: ResetPasswordRequest):
    """Şifreyi sıfırla"""
    # Token kontrolü
    reset_record = await db.password_resets.find_one({
        "token": input.token,
        "used": False
    })
    
    if not reset_record:
        raise HTTPException(status_code=400, detail="Geçersiz veya süresi dolmuş token")
    
    # Süre kontrolü
    expires_at = datetime.fromisoformat(reset_record["expires_at"].replace('Z', '+00:00'))
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="Token süresi dolmuş")
    
    # Şifre validasyonu
    if len(input.new_password) < 6:
        raise HTTPException(status_code=400, detail="Şifre en az 6 karakter olmalı")
    
    # Şifreyi güncelle
    new_hash = hash_password(input.new_password)
    await db.users.update_one(
        {"id": reset_record["user_id"]},
        {"$set": {"password_hash": new_hash}}
    )
    
    # Token'ı kullanılmış olarak işaretle
    await db.password_resets.update_one(
        {"token": input.token},
        {"$set": {"used": True}}
    )
    
    return {"message": "Şifreniz başarıyla güncellendi"}

# =============================================================================
# Customer endpoints (Mevcut - değiştirilmedi)
# =============================================================================
@api_router.get("/")
async def root():
    return {"message": "Müşteri Ziyaret Takip API"}

# Region endpoints - FAZ 3.2: user_id filtresi eklendi
@api_router.get("/regions", response_model=List[Region])
async def get_regions(current_user: dict = Depends(require_auth)):
    """Kullanıcının bölgelerini listele"""
    regions = await db.regions.find({"user_id": current_user["id"]}, {"_id": 0}).to_list(1000)
    for r in regions:
        if isinstance(r.get('created_at'), str):
            r['created_at'] = datetime.fromisoformat(r['created_at'])
    return regions

@api_router.get("/regions/{region_id}")
async def get_region(region_id: str, current_user: dict = Depends(require_auth)):
    """Kullanıcının bölge detayını getir"""
    region = await db.regions.find_one({"id": region_id, "user_id": current_user["id"]}, {"_id": 0})
    if not region:
        raise HTTPException(status_code=404, detail="Bölge bulunamadı")
    
    # Get customer count (only user's customers)
    customer_count = await db.customers.count_documents({"region": region["name"], "user_id": current_user["id"]})
    region["customer_count"] = customer_count
    
    if isinstance(region.get('created_at'), str):
        region['created_at'] = datetime.fromisoformat(region['created_at'])
    return region

@api_router.post("/regions", response_model=Region)
async def create_region(input: RegionCreate, current_user: dict = Depends(require_auth)):
    """Yeni bölge oluştur"""
    # Check if region name already exists for this user
    existing = await db.regions.find_one({"name": input.name, "user_id": current_user["id"]})
    if existing:
        raise HTTPException(status_code=400, detail="Bu isimde bir bölge zaten var")
    
    region_obj = Region(**input.model_dump(), user_id=current_user["id"])
    doc = region_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.regions.insert_one(doc)
    return region_obj

@api_router.put("/regions/{region_id}", response_model=Region)
async def update_region(region_id: str, input: RegionUpdate, current_user: dict = Depends(require_auth)):
    """Bölge güncelle"""
    region = await db.regions.find_one({"id": region_id, "user_id": current_user["id"]}, {"_id": 0})
    if not region:
        raise HTTPException(status_code=404, detail="Bölge bulunamadı")
    
    old_name = region["name"]
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    
    # Check if new name already exists for this user
    if "name" in update_data and update_data["name"] != old_name:
        existing = await db.regions.find_one({"name": update_data["name"], "user_id": current_user["id"]})
        if existing:
            raise HTTPException(status_code=400, detail="Bu isimde bir bölge zaten var")
    
    if update_data:
        await db.regions.update_one({"id": region_id, "user_id": current_user["id"]}, {"$set": update_data})
        
        # Update customer regions if name changed (only user's customers)
        if "name" in update_data and update_data["name"] != old_name:
            await db.customers.update_many(
                {"region": old_name, "user_id": current_user["id"]},
                {"$set": {"region": update_data["name"]}}
            )
    
    updated = await db.regions.find_one({"id": region_id, "user_id": current_user["id"]}, {"_id": 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return updated

@api_router.delete("/regions/{region_id}")
async def delete_region(region_id: str, current_user: dict = Depends(require_auth)):
    """Bölge sil"""
    region = await db.regions.find_one({"id": region_id, "user_id": current_user["id"]}, {"_id": 0})
    if not region:
        raise HTTPException(status_code=404, detail="Bölge bulunamadı")
    
    # Check if there are customers in this region (only user's customers)
    customer_count = await db.customers.count_documents({"region": region["name"], "user_id": current_user["id"]})
    if customer_count > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Bu bölgede {customer_count} müşteri var. Önce müşterileri başka bölgeye taşıyın."
        )
    
    await db.regions.delete_one({"id": region_id, "user_id": current_user["id"]})
    return {"message": "Bölge silindi"}

@api_router.get("/regions/{region_id}/customers")
async def get_region_customers(region_id: str, current_user: dict = Depends(require_auth)):
    """Bölgedeki müşterileri getir"""
    region = await db.regions.find_one({"id": region_id, "user_id": current_user["id"]}, {"_id": 0})
    if not region:
        raise HTTPException(status_code=404, detail="Bölge bulunamadı")
    
    customers = await db.customers.find({"region": region["name"], "user_id": current_user["id"]}, {"_id": 0}).to_list(1000)
    for c in customers:
        if isinstance(c.get('created_at'), str):
            c['created_at'] = datetime.fromisoformat(c['created_at'])
    return customers

# Customer endpoints - FAZ 3.2: user_id filtresi eklendi
@api_router.get("/customers", response_model=List[Customer])
async def get_customers(current_user: dict = Depends(require_auth)):
    """Kullanıcının müşterilerini listele"""
    customers = await db.customers.find({"user_id": current_user["id"]}, {"_id": 0}).to_list(1000)
    for c in customers:
        if isinstance(c.get('created_at'), str):
            c['created_at'] = datetime.fromisoformat(c['created_at'])
    return customers

# Download sample Excel template - MUST be before /{customer_id} route
@api_router.get("/customers/template")
async def download_template():
    """Örnek Excel şablonu indir"""
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Müşteriler"
    
    # Headers
    headers = ["Müşteri Adı", "Bölge", "Telefon", "Adres", "Fiyat Statüsü", "Ziyaret Günleri"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Sample data
    sample_data = [
        ["Örnek Market", "Kadıköy", "0532 111 2233", "Moda Cad. No:15", "Standart", "Pazartesi, Çarşamba, Cuma"],
        ["ABC Bakkaliye", "Beşiktaş", "0533 222 3344", "Çarşı Cad. No:8", "İskontolu", "Salı, Perşembe"],
    ]
    for row_num, row_data in enumerate(sample_data, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=value)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 35
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=musteri_sablonu.xlsx"}
    )

@api_router.get("/customers/{customer_id}", response_model=Customer)
async def get_customer(customer_id: str, current_user: dict = Depends(require_auth)):
    """Müşteri detayını getir"""
    customer = await db.customers.find_one({"id": customer_id, "user_id": current_user["id"]}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    if isinstance(customer.get('created_at'), str):
        customer['created_at'] = datetime.fromisoformat(customer['created_at'])
    return customer

@api_router.post("/customers", response_model=Customer)
async def create_customer(input: CustomerCreate, current_user: dict = Depends(require_auth)):
    """Yeni müşteri oluştur"""
    customer_obj = Customer(**input.model_dump(), user_id=current_user["id"])
    doc = customer_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.customers.insert_one(doc)
    return customer_obj

@api_router.put("/customers/{customer_id}", response_model=Customer)
async def update_customer(customer_id: str, input: CustomerUpdate, current_user: dict = Depends(require_auth)):
    """Müşteri güncelle"""
    customer = await db.customers.find_one({"id": customer_id, "user_id": current_user["id"]}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    if update_data:
        await db.customers.update_one({"id": customer_id, "user_id": current_user["id"]}, {"$set": update_data})
    
    updated = await db.customers.find_one({"id": customer_id, "user_id": current_user["id"]}, {"_id": 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return updated

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, current_user: dict = Depends(require_auth)):
    """Müşteri sil"""
    # Önce müşterinin bu kullanıcıya ait olduğunu doğrula
    customer = await db.customers.find_one({"id": customer_id, "user_id": current_user["id"]}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    
    await db.customers.delete_one({"id": customer_id, "user_id": current_user["id"]})
    # Delete related visits and follow-ups (only user's data)
    await db.visits.delete_many({"customer_id": customer_id, "user_id": current_user["id"]})
    await db.follow_ups.delete_many({"customer_id": customer_id, "user_id": current_user["id"]})
    return {"message": "Müşteri silindi"}

# Follow-Up endpoints - FAZ 3.2: user_id filtresi eklendi
@api_router.get("/follow-ups")
async def get_follow_ups(
    date: Optional[str] = None, 
    customer_id: Optional[str] = None, 
    status: Optional[str] = None,
    current_user: dict = Depends(require_auth)
):
    """Kullanıcının takiplerini listele"""
    query = {"user_id": current_user["id"]}
    if date:
        query["due_date"] = date
    if customer_id:
        query["customer_id"] = customer_id
    if status:
        query["status"] = status
    
    follow_ups = await db.follow_ups.find(query, {"_id": 0}).to_list(1000)
    
    # Update late status for overdue follow-ups
    today = datetime.now(timezone.utc).date().isoformat()
    for fu in follow_ups:
        if fu.get("status") == "pending" and fu.get("due_date") < today:
            fu["status"] = "late"
            await db.follow_ups.update_one({"id": fu["id"]}, {"$set": {"status": "late"}})
    
    return follow_ups

@api_router.get("/follow-ups/today")
async def get_today_follow_ups(current_user: dict = Depends(require_auth)):
    """Bugünkü ve gecikmiş takipleri getir"""
    today = datetime.now(timezone.utc).date().isoformat()
    
    # Get today's and overdue follow-ups (only user's)
    follow_ups = await db.follow_ups.find({
        "user_id": current_user["id"],
        "$or": [
            {"due_date": today},
            {"due_date": {"$lt": today}, "status": {"$ne": "done"}}
        ]
    }, {"_id": 0}).to_list(1000)
    
    # Update late status
    for fu in follow_ups:
        if fu.get("status") == "pending" and fu.get("due_date") < today:
            fu["status"] = "late"
            await db.follow_ups.update_one({"id": fu["id"]}, {"$set": {"status": "late"}})
    
    # Get customer info for each follow-up (only user's customers)
    result = []
    for fu in follow_ups:
        customer = await db.customers.find_one({"id": fu["customer_id"], "user_id": current_user["id"]}, {"_id": 0})
        if customer:
            fu["customer"] = {"name": customer["name"], "region": customer["region"]}
        result.append(fu)
    
    return result

@api_router.get("/follow-ups/{follow_up_id}")
async def get_follow_up(follow_up_id: str, current_user: dict = Depends(require_auth)):
    """Takip detayını getir"""
    fu = await db.follow_ups.find_one({"id": follow_up_id, "user_id": current_user["id"]}, {"_id": 0})
    if not fu:
        raise HTTPException(status_code=404, detail="Takip bulunamadı")
    return fu

@api_router.post("/follow-ups", response_model=FollowUp)
async def create_follow_up(input: FollowUpCreate, current_user: dict = Depends(require_auth)):
    """Yeni takip oluştur"""
    # Verify customer exists and belongs to user
    customer = await db.customers.find_one({"id": input.customer_id, "user_id": current_user["id"]})
    if not customer:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    
    fu_obj = FollowUp(**input.model_dump(), user_id=current_user["id"])
    doc = fu_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.follow_ups.insert_one(doc)
    return fu_obj

@api_router.put("/follow-ups/{follow_up_id}")
async def update_follow_up(follow_up_id: str, input: FollowUpUpdate, current_user: dict = Depends(require_auth)):
    """Takip güncelle"""
    fu = await db.follow_ups.find_one({"id": follow_up_id, "user_id": current_user["id"]}, {"_id": 0})
    if not fu:
        raise HTTPException(status_code=404, detail="Takip bulunamadı")
    
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    
    # If marking as done, set completed_at
    if update_data.get("status") == "done":
        update_data["completed_at"] = datetime.now(timezone.utc).isoformat()
    
    if update_data:
        await db.follow_ups.update_one({"id": follow_up_id, "user_id": current_user["id"]}, {"$set": update_data})
    
    updated = await db.follow_ups.find_one({"id": follow_up_id, "user_id": current_user["id"]}, {"_id": 0})
    return updated

@api_router.delete("/follow-ups/{follow_up_id}")
async def delete_follow_up(follow_up_id: str, current_user: dict = Depends(require_auth)):
    """Takip sil"""
    # Önce takibin bu kullanıcıya ait olduğunu doğrula
    fu = await db.follow_ups.find_one({"id": follow_up_id, "user_id": current_user["id"]}, {"_id": 0})
    if not fu:
        raise HTTPException(status_code=404, detail="Takip bulunamadı")
    
    await db.follow_ups.delete_one({"id": follow_up_id, "user_id": current_user["id"]})
    return {"message": "Takip silindi"}

@api_router.post("/follow-ups/{follow_up_id}/complete")
async def complete_follow_up(follow_up_id: str, current_user: dict = Depends(require_auth)):
    """Takibi tamamla"""
    fu = await db.follow_ups.find_one({"id": follow_up_id, "user_id": current_user["id"]}, {"_id": 0})
    if not fu:
        raise HTTPException(status_code=404, detail="Takip bulunamadı")
    
    await db.follow_ups.update_one(
        {"id": follow_up_id, "user_id": current_user["id"]}, 
        {"$set": {"status": "done", "completed_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": "Takip tamamlandı"}

# Get customers for today based on visit_days - FAZ 3.2: user_id filtresi eklendi
@api_router.get("/customers/today/{day_name}")
async def get_today_customers(day_name: str, current_user: dict = Depends(require_auth)):
    """Bugün ziyaret edilecek müşterileri getir"""
    customers = await db.customers.find(
        {"visit_days": day_name, "user_id": current_user["id"]}, 
        {"_id": 0}
    ).to_list(1000)
    for c in customers:
        if isinstance(c.get('created_at'), str):
            c['created_at'] = datetime.fromisoformat(c['created_at'])
    return customers

# Visit endpoints - FAZ 3.2: user_id filtresi eklendi
# Helper function: Eski verilere status alanı ekle (geriye uyumluluk)
def migrate_visit_status(visit: dict) -> dict:
    """Eski visit kayıtlarına status alanı ekle"""
    if "status" not in visit or visit.get("status") is None:
        if visit.get("completed"):
            visit["status"] = "visited"
        elif visit.get("visit_skip_reason"):
            visit["status"] = "not_visited"
        else:
            visit["status"] = "pending"
    return visit

@api_router.get("/visits", response_model=List[Visit])
async def get_visits(
    date: Optional[str] = None, 
    customer_id: Optional[str] = None,
    current_user: dict = Depends(require_auth)
):
    """Kullanıcının ziyaretlerini listele"""
    query = {"user_id": current_user["id"]}
    if date:
        query["date"] = date
    if customer_id:
        query["customer_id"] = customer_id
    
    visits = await db.visits.find(query, {"_id": 0}).to_list(1000)
    for v in visits:
        # Geriye uyumluluk: status alanı ekle
        migrate_visit_status(v)
        if isinstance(v.get('created_at'), str):
            v['created_at'] = datetime.fromisoformat(v['created_at'])
        if isinstance(v.get('completed_at'), str):
            v['completed_at'] = datetime.fromisoformat(v['completed_at'])
    return visits

@api_router.get("/visits/{visit_id}", response_model=Visit)
async def get_visit(visit_id: str, current_user: dict = Depends(require_auth)):
    """Ziyaret detayını getir"""
    visit = await db.visits.find_one({"id": visit_id, "user_id": current_user["id"]}, {"_id": 0})
    if not visit:
        raise HTTPException(status_code=404, detail="Ziyaret bulunamadı")
    # Geriye uyumluluk: status alanı ekle
    migrate_visit_status(visit)
    if isinstance(visit.get('created_at'), str):
        visit['created_at'] = datetime.fromisoformat(visit['created_at'])
    if isinstance(visit.get('completed_at'), str):
        visit['completed_at'] = datetime.fromisoformat(visit['completed_at'])
    return visit

@api_router.post("/visits", response_model=Visit)
async def create_or_get_visit(customer_id: str, date: str, current_user: dict = Depends(require_auth)):
    """Ziyaret oluştur veya mevcut ziyareti getir"""
    # Müşterinin bu kullanıcıya ait olduğunu doğrula
    customer = await db.customers.find_one({"id": customer_id, "user_id": current_user["id"]}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    
    # Check if visit already exists for this user
    existing = await db.visits.find_one({
        "customer_id": customer_id, 
        "date": date, 
        "user_id": current_user["id"]
    }, {"_id": 0})
    if existing:
        # Geriye uyumluluk: status alanı ekle
        migrate_visit_status(existing)
        if isinstance(existing.get('created_at'), str):
            existing['created_at'] = datetime.fromisoformat(existing['created_at'])
        if isinstance(existing.get('completed_at'), str):
            existing['completed_at'] = datetime.fromisoformat(existing['completed_at'])
        return existing
    
    # Create new visit with status=pending
    visit_obj = Visit(customer_id=customer_id, date=date, user_id=current_user["id"], status="pending")
    doc = visit_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    if doc.get('completed_at'):
        doc['completed_at'] = doc['completed_at'].isoformat()
    await db.visits.insert_one(doc)
    return visit_obj

@api_router.put("/visits/{visit_id}", response_model=Visit)
async def update_visit(visit_id: str, input: VisitUpdate, current_user: dict = Depends(require_auth)):
    """Ziyaret güncelle"""
    visit = await db.visits.find_one({"id": visit_id, "user_id": current_user["id"]}, {"_id": 0})
    if not visit:
        raise HTTPException(status_code=404, detail="Ziyaret bulunamadı")
    
    update_data = {}
    for k, v in input.model_dump().items():
        if v is not None:
            update_data[k] = v
        elif k in ['visit_skip_reason', 'payment_skip_reason', 'payment_type', 'payment_amount', 'customer_request', 'note', 'quality_rating']:
            # Allow clearing these fields
            if input.model_dump(exclude_unset=True).get(k) is not None or k in input.model_dump(exclude_unset=True):
                update_data[k] = v
    
    # Status değişikliğine göre completed alanını da güncelle (geriye uyumluluk)
    if 'status' in update_data:
        if update_data['status'] == 'visited':
            update_data['completed'] = True
            update_data['completed_at'] = datetime.now(timezone.utc).isoformat()
            update_data['visit_skip_reason'] = None
        elif update_data['status'] == 'not_visited':
            update_data['completed'] = False
            update_data['completed_at'] = None
            # payment alanlarını temizle
            update_data['payment_collected'] = False
            update_data['payment_type'] = None
            update_data['payment_amount'] = None
            update_data['payment_skip_reason'] = None
        elif update_data['status'] == 'pending':
            update_data['completed'] = False
            update_data['completed_at'] = None
            update_data['visit_skip_reason'] = None
    # Eski completed tabanlı güncelleme için de status'u senkronize et
    elif 'completed' in update_data:
        if update_data['completed']:
            update_data['status'] = 'visited'
            update_data['completed_at'] = datetime.now(timezone.utc).isoformat()
        else:
            # visit_skip_reason varsa not_visited, yoksa pending
            if update_data.get('visit_skip_reason') or visit.get('visit_skip_reason'):
                update_data['status'] = 'not_visited'
            else:
                update_data['status'] = 'pending'
    
    if update_data:
        await db.visits.update_one({"id": visit_id, "user_id": current_user["id"]}, {"$set": update_data})
    
    updated = await db.visits.find_one({"id": visit_id, "user_id": current_user["id"]}, {"_id": 0})
    # Geriye uyumluluk: status alanı ekle
    migrate_visit_status(updated)
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    if isinstance(updated.get('completed_at'), str):
        updated['completed_at'] = datetime.fromisoformat(updated['completed_at'])
    if isinstance(updated.get('started_at'), str):
        updated['started_at'] = datetime.fromisoformat(updated['started_at'])
    if isinstance(updated.get('ended_at'), str):
        updated['ended_at'] = datetime.fromisoformat(updated['ended_at'])
    return updated

# FAZ 2: Ziyaret Süresi Takibi Endpoint'leri - FAZ 3.2: user_id filtresi eklendi
@api_router.post("/visits/{visit_id}/start")
async def start_visit(visit_id: str, current_user: dict = Depends(require_auth)):
    """Ziyareti başlat - started_at zamanını kaydet"""
    visit = await db.visits.find_one({"id": visit_id, "user_id": current_user["id"]}, {"_id": 0})
    if not visit:
        raise HTTPException(status_code=404, detail="Ziyaret bulunamadı")
    
    if visit.get("started_at"):
        raise HTTPException(status_code=400, detail="Ziyaret zaten başlatılmış")
    
    now = datetime.now(timezone.utc)
    await db.visits.update_one(
        {"id": visit_id, "user_id": current_user["id"]}, 
        {"$set": {"started_at": now.isoformat()}}
    )
    
    return {"message": "Ziyaret başlatıldı", "started_at": now.isoformat()}

@api_router.post("/visits/{visit_id}/end")
async def end_visit(visit_id: str, current_user: dict = Depends(require_auth)):
    """Ziyareti bitir - ended_at zamanını kaydet ve süreyi hesapla"""
    visit = await db.visits.find_one({"id": visit_id, "user_id": current_user["id"]}, {"_id": 0})
    if not visit:
        raise HTTPException(status_code=404, detail="Ziyaret bulunamadı")
    
    if not visit.get("started_at"):
        raise HTTPException(status_code=400, detail="Ziyaret henüz başlatılmamış")
    
    if visit.get("ended_at"):
        raise HTTPException(status_code=400, detail="Ziyaret zaten bitirilmiş")
    
    now = datetime.now(timezone.utc)
    started_at = visit.get("started_at")
    if isinstance(started_at, str):
        started_at = datetime.fromisoformat(started_at.replace('Z', '+00:00'))
    
    # Süreyi dakika olarak hesapla
    duration = int((now - started_at).total_seconds() / 60)
    
    await db.visits.update_one(
        {"id": visit_id, "user_id": current_user["id"]}, 
        {"$set": {
            "ended_at": now.isoformat(),
            "duration_minutes": duration
        }}
    )
    
    return {
        "message": "Ziyaret tamamlandı", 
        "ended_at": now.isoformat(),
        "duration_minutes": duration
    }

# Müşteri uyarı seçenekleri endpoint'i
@api_router.get("/customer-alerts")
async def get_customer_alert_options():
    """Müşteri uyarı seçeneklerini döndür"""
    return {"alerts": CUSTOMER_ALERTS}

# Daily Report Note endpoints - FAZ 3.2: user_id filtresi eklendi
@api_router.get("/daily-note/{date}")
async def get_daily_note(date: str, current_user: dict = Depends(require_auth)):
    """Gün sonu notunu getir"""
    note = await db.daily_notes.find_one({"date": date, "user_id": current_user["id"]}, {"_id": 0})
    if not note:
        return {"date": date, "note": ""}
    return note

@api_router.post("/daily-note/{date}")
async def save_daily_note(date: str, input: DailyReportNoteUpdate, current_user: dict = Depends(require_auth)):
    """Gün sonu notunu kaydet"""
    existing = await db.daily_notes.find_one({"date": date, "user_id": current_user["id"]})
    if existing:
        await db.daily_notes.update_one(
            {"date": date, "user_id": current_user["id"]}, 
            {"$set": {"note": input.note}}
        )
    else:
        note_obj = DailyReportNote(date=date, note=input.note)
        doc = note_obj.model_dump()
        doc['user_id'] = current_user["id"]
        doc['created_at'] = doc['created_at'].isoformat()
        await db.daily_notes.insert_one(doc)
    return {"message": "Not kaydedildi", "date": date}

# Excel Upload endpoint - FAZ 3.2: user_id filtresi eklendi
@api_router.post("/customers/upload")
async def upload_customers_excel(file: UploadFile = File(...), current_user: dict = Depends(require_auth)):
    """
    Excel dosyasından toplu müşteri yükleme.
    Gerekli sütunlar: Müşteri Adı, Bölge
    Opsiyonel sütunlar: Telefon, Adres, Fiyat Statüsü, Ziyaret Günleri
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Sadece Excel dosyaları (.xlsx, .xls) kabul edilir")
    
    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents))
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip().lower() if cell.value else "")
        
        # Map column names (Turkish to English)
        column_map = {
            'müşteri adı': 'name',
            'musteri adi': 'name',
            'ad': 'name',
            'isim': 'name',
            'name': 'name',
            'bölge': 'region',
            'bolge': 'region',
            'region': 'region',
            'telefon': 'phone',
            'tel': 'phone',
            'phone': 'phone',
            'adres': 'address',
            'address': 'address',
            'fiyat statüsü': 'price_status',
            'fiyat statusu': 'price_status',
            'fiyat': 'price_status',
            'statü': 'price_status',
            'price_status': 'price_status',
            'ziyaret günleri': 'visit_days',
            'ziyaret gunleri': 'visit_days',
            'günler': 'visit_days',
            'visit_days': 'visit_days',
        }
        
        # Find column indices
        col_indices = {}
        for i, header in enumerate(headers):
            mapped = column_map.get(header)
            if mapped:
                col_indices[mapped] = i
        
        # Check required columns
        if 'name' not in col_indices:
            raise HTTPException(status_code=400, detail="'Müşteri Adı' sütunu bulunamadı")
        if 'region' not in col_indices:
            raise HTTPException(status_code=400, detail="'Bölge' sütunu bulunamadı")
        
        customers_to_add = []
        errors = []
        row_num = 1
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            
            # Skip empty rows
            if not any(row):
                continue
            
            name = str(row[col_indices['name']]).strip() if row[col_indices['name']] else ""
            region = str(row[col_indices['region']]).strip() if row[col_indices['region']] else ""
            
            if not name or not region:
                errors.append(f"Satır {row_num}: Müşteri adı veya bölge boş")
                continue
            
            customer = {
                "id": str(uuid.uuid4()),
                "name": name,
                "region": region,
                "phone": str(row[col_indices.get('phone', -1)]).strip() if col_indices.get('phone') is not None and row[col_indices['phone']] else None,
                "address": str(row[col_indices.get('address', -1)]).strip() if col_indices.get('address') is not None and row[col_indices['address']] else None,
                "price_status": "Standart",
                "visit_days": [],
                "alerts": [],
                "user_id": current_user["id"],  # FAZ 3.2: user_id eklendi
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            # Handle price_status
            if 'price_status' in col_indices and row[col_indices['price_status']]:
                ps = str(row[col_indices['price_status']]).strip()
                ps_lower = ps.lower().replace('i̇', 'i').replace('İ', 'i')
                if ps_lower in ['iskontolu', 'iskonto', 'indirimli', 'özel'] or ps == "İskontolu":
                    customer['price_status'] = "İskontolu"
            
            # Handle visit_days
            if 'visit_days' in col_indices and row[col_indices['visit_days']]:
                days_str = str(row[col_indices['visit_days']]).strip()
                valid_days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
                customer['visit_days'] = [d.strip() for d in days_str.split(',') if d.strip() in valid_days]
            
            customers_to_add.append(customer)
        
        if not customers_to_add:
            raise HTTPException(status_code=400, detail="Yüklenecek geçerli müşteri bulunamadı")
        
        # Insert customers
        await db.customers.insert_many(customers_to_add)
        
        return {
            "message": f"{len(customers_to_add)} müşteri başarıyla yüklendi",
            "added_count": len(customers_to_add),
            "errors": errors[:10] if errors else []  # Return first 10 errors
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Dosya işlenirken hata: {str(e)}")

# Analytics endpoints - FAZ 3.2: user_id filtresi eklendi
@api_router.get("/analytics/performance")
async def get_performance_analytics(
    period: str = "weekly", 
    start_date: str = None, 
    end_date: str = None,
    current_user: dict = Depends(require_auth)
):
    """
    Get performance analytics for a given period.
    period: 'weekly' or 'monthly'
    """
    from datetime import timedelta
    
    today = datetime.now(timezone.utc).date()
    
    if period == "weekly":
        # Current week (Monday to Sunday)
        days_since_monday = today.weekday()
        week_start = today - timedelta(days=days_since_monday)
        week_end = week_start + timedelta(days=6)
        start = week_start.isoformat()
        end = week_end.isoformat()
    else:  # monthly
        # Current month
        start = today.replace(day=1).isoformat()
        # Last day of month
        if today.month == 12:
            next_month = today.replace(year=today.year + 1, month=1, day=1)
        else:
            next_month = today.replace(month=today.month + 1, day=1)
        end = (next_month - timedelta(days=1)).isoformat()
    
    # Override with custom dates if provided
    if start_date:
        start = start_date
    if end_date:
        end = end_date
    
    # Get all customers (only user's customers)
    all_customers = await db.customers.find({"user_id": current_user["id"]}, {"_id": 0}).to_list(1000)
    
    # Get visits in date range (only user's visits)
    visits = await db.visits.find({
        "user_id": current_user["id"],
        "date": {"$gte": start, "$lte": end}
    }, {"_id": 0}).to_list(10000)
    
    # Get follow-ups in date range for planned visit calculation (only user's)
    follow_ups = await db.follow_ups.find({
        "user_id": current_user["id"],
        "due_date": {"$gte": start, "$lte": end}
    }, {"_id": 0}).to_list(10000)
    
    # Calculate metrics based on follow-ups
    total_planned = len(follow_ups)  # Planlanan ziyaret = Toplam takip sayısı
    total_completed = sum(1 for fu in follow_ups if fu.get("status") == "done")  # Tamamlanan takipler
    
    total_payment = 0
    payment_count = 0
    payment_skip_reasons = {}
    visit_skip_reasons = {}
    
    end_dt = datetime.fromisoformat(end)
    
    # Process visits for payment data
    visits_by_customer = {}
    visit_completed_count = 0  # Ziyaret tamamlama sayısı (ödeme oranı için)
    
    # FAZ 2: Ziyaret süresi ve kalite metrikleri
    duration_values = []
    quality_values = []
    quality_payment_data = []  # (kalite, tahsilat) tuple'ları
    
    for v in visits:
        cid = v.get("customer_id")
        if cid not in visits_by_customer:
            visits_by_customer[cid] = []
        visits_by_customer[cid].append(v)
        
        if v.get("completed"):
            visit_completed_count += 1
        elif v.get("visit_skip_reason"):
            reason = v.get("visit_skip_reason", "Belirtilmemiş")
            visit_skip_reasons[reason] = visit_skip_reasons.get(reason, 0) + 1
        
        if v.get("payment_collected"):
            payment_count += 1
            payment_amount = v.get("payment_amount", 0) or 0
            total_payment += payment_amount
            
            # Kalite-Tahsilat ilişkisi için veri topla
            if v.get("quality_rating"):
                quality_payment_data.append((v.get("quality_rating"), payment_amount))
        elif v.get("payment_skip_reason"):
            reason = v.get("payment_skip_reason", "Belirtilmemiş")
            payment_skip_reasons[reason] = payment_skip_reasons.get(reason, 0) + 1
        
        # FAZ 2: Süre ve kalite verilerini topla
        if v.get("duration_minutes") is not None:
            duration_values.append(v.get("duration_minutes"))
        if v.get("quality_rating") is not None:
            quality_values.append(v.get("quality_rating"))
    
    # FAZ 2: Süre analizi
    avg_duration = round(sum(duration_values) / len(duration_values), 1) if duration_values else None
    short_visits = len([d for d in duration_values if d < 5])  # <5 dakika
    long_visits = len([d for d in duration_values if d > 60])  # >60 dakika
    
    # FAZ 2: Kalite analizi
    avg_quality = round(sum(quality_values) / len(quality_values), 1) if quality_values else None
    quality_distribution = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    for q in quality_values:
        if q in quality_distribution:
            quality_distribution[q] += 1
    
    # Kalite-Tahsilat ilişkisi (ortalama tahsilat her kalite seviyesi için)
    quality_payment_relation = {}
    for rating in range(1, 6):
        payments = [p for (q, p) in quality_payment_data if q == rating]
        if payments:
            quality_payment_relation[rating] = round(sum(payments) / len(payments), 2)
    
    # New customers in period
    new_customers = []
    for c in all_customers:
        created = c.get("created_at", "")
        if isinstance(created, str):
            created_date = created[:10]
            if start <= created_date <= end:
                new_customers.append(c)
    
    # Price status analysis
    iskontolu_customers = [c for c in all_customers if c.get("price_status") == "İskontolu"]
    standart_customers = [c for c in all_customers if c.get("price_status") != "İskontolu"]
    
    # Visits and payments by price status
    iskontolu_visits = 0
    iskontolu_completed = 0
    iskontolu_payment = 0
    standart_visits = 0
    standart_completed = 0
    standart_payment = 0
    
    customer_map = {c["id"]: c for c in all_customers}
    for v in visits:
        cid = v.get("customer_id")
        customer = customer_map.get(cid, {})
        is_iskontolu = customer.get("price_status") == "İskontolu"
        
        if is_iskontolu:
            iskontolu_visits += 1
            if v.get("completed"):
                iskontolu_completed += 1
            if v.get("payment_collected"):
                iskontolu_payment += v.get("payment_amount", 0) or 0
        else:
            standart_visits += 1
            if v.get("completed"):
                standart_completed += 1
            if v.get("payment_collected"):
                standart_payment += v.get("payment_amount", 0) or 0
    
    # Daily breakdown for charts
    daily_data = []
    date_cursor = datetime.fromisoformat(start)
    while date_cursor.date() <= end_dt.date():
        date_str = date_cursor.date().isoformat()
        day_visits = [v for v in visits if v.get("date") == date_str]
        day_followups = [fu for fu in follow_ups if fu.get("due_date") == date_str]
        completed_followups = sum(1 for fu in day_followups if fu.get("status") == "done")
        payment = sum((v.get("payment_amount", 0) or 0) for v in day_visits if v.get("payment_collected"))
        
        daily_data.append({
            "date": date_str,
            "day": date_cursor.strftime("%a"),
            "planned": len(day_followups),  # Planlanan = O günkü takipler
            "completed": completed_followups,  # Tamamlanan = O günkü tamamlanan takipler
            "payment": payment
        })
        date_cursor += timedelta(days=1)
    
    visit_rate = (total_completed / total_planned * 100) if total_planned > 0 else 0
    payment_rate = (payment_count / visit_completed_count * 100) if visit_completed_count > 0 else 0
    
    return {
        "period": period,
        "start_date": start,
        "end_date": end,
        "visit_performance": {
            "total_planned": total_planned,
            "total_completed": total_completed,
            "visit_rate": round(visit_rate, 1),
            "skip_reasons": visit_skip_reasons
        },
        "payment_performance": {
            "total_amount": total_payment,
            "customer_count": payment_count,
            "payment_rate": round(payment_rate, 1),
            "skip_reasons": payment_skip_reasons
        },
        "customer_acquisition": {
            "new_count": len(new_customers),
            "new_customers": [{"name": c["name"], "region": c["region"], "price_status": c.get("price_status", "Standart")} for c in new_customers]
        },
        "price_analysis": {
            "iskontolu": {
                "customer_count": len(iskontolu_customers),
                "visit_count": iskontolu_visits,
                "completed_count": iskontolu_completed,
                "visit_rate": round((iskontolu_completed / iskontolu_visits * 100) if iskontolu_visits > 0 else 0, 1),
                "total_payment": iskontolu_payment
            },
            "standart": {
                "customer_count": len(standart_customers),
                "visit_count": standart_visits,
                "completed_count": standart_completed,
                "visit_rate": round((standart_completed / standart_visits * 100) if standart_visits > 0 else 0, 1),
                "total_payment": standart_payment
            }
        },
        "daily_breakdown": daily_data,
        # FAZ 2: Ziyaret Kalitesi Metrikleri
        "visit_quality": {
            "duration": {
                "average_minutes": avg_duration,
                "total_measured": len(duration_values),
                "short_visits": short_visits,  # <5 dakika
                "long_visits": long_visits,  # >60 dakika
                "warning_threshold": {"short": 5, "long": 60}
            },
            "rating": {
                "average_rating": avg_quality,
                "total_rated": len(quality_values),
                "distribution": quality_distribution,
                "quality_payment_relation": quality_payment_relation
            }
        }
    }

# Seed sample data
@api_router.post("/seed")
async def seed_data():
    # Check if data already exists
    count = await db.customers.count_documents({})
    if count > 0:
        return {"message": "Veriler zaten mevcut", "customer_count": count}
    
    sample_customers = [
        {
            "id": str(uuid.uuid4()),
            "name": "Ahmet Yılmaz Market",
            "region": "Kadıköy",
            "phone": "0532 111 2233",
            "address": "Caferağa Mah. Moda Cad. No:15",
            "price_status": "İskontolu",
            "visit_days": ["Pazartesi", "Perşembe"],
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Elif Bakkaliye",
            "region": "Beşiktaş",
            "phone": "0533 222 3344",
            "address": "Sinanpaşa Mah. Çarşı Cad. No:8",
            "price_status": "Standart",
            "visit_days": ["Salı", "Cuma"],
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Mehmet Gıda",
            "region": "Şişli",
            "phone": "0534 333 4455",
            "address": "Meşrutiyet Mah. Halaskargazi Cad. No:42",
            "price_status": "İskontolu",
            "visit_days": ["Pazartesi", "Çarşamba", "Cuma"],
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Ayşe Manav",
            "region": "Üsküdar",
            "phone": "0535 444 5566",
            "address": "Altunizade Mah. Kısıklı Cad. No:23",
            "price_status": "Standart",
            "visit_days": ["Salı", "Perşembe", "Cumartesi"],
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Can Süpermarket",
            "region": "Kadıköy",
            "phone": "0536 555 6677",
            "address": "Fenerbahçe Mah. Bağdat Cad. No:156",
            "price_status": "Standart",
            "visit_days": ["Çarşamba", "Cumartesi"],
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Demir Ticaret",
            "region": "Maltepe",
            "phone": "0537 666 7788",
            "address": "Cevizli Mah. D-100 Yan Yol No:88",
            "price_status": "İskontolu",
            "visit_days": ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"],
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Güneş Market",
            "region": "Ataşehir",
            "phone": "0538 777 8899",
            "address": "İçerenköy Mah. Kayışdağı Cad. No:34",
            "price_status": "Standart",
            "visit_days": ["Perşembe", "Pazar"],
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Yıldız Bakkal",
            "region": "Beşiktaş",
            "phone": "0539 888 9900",
            "address": "Levent Mah. Nispetiye Cad. No:67",
            "price_status": "İskontolu",
            "visit_days": ["Pazartesi", "Cuma"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
    ]
    
    await db.customers.insert_many(sample_customers)
    return {"message": "Örnek veriler eklendi", "customer_count": len(sample_customers)}

# =============================================================================
# FAZ 4: Araç, Yakıt ve Günlük KM Takibi Endpoint'leri
# =============================================================================

# Yakıt türleri listesi
@api_router.get("/fuel-types")
async def get_fuel_types():
    """Desteklenen yakıt türlerini döndür"""
    return {"fuel_types": FUEL_TYPES}

# ===== ARAÇ YÖNETİMİ =====
@api_router.get("/vehicles")
async def get_vehicles(current_user: dict = Depends(require_auth)):
    """Kullanıcının araçlarını listele"""
    vehicles = await db.vehicles.find(
        {"user_id": current_user["id"]}, 
        {"_id": 0}
    ).to_list(100)
    return vehicles

@api_router.get("/vehicles/active")
async def get_active_vehicle(current_user: dict = Depends(require_auth)):
    """Kullanıcının aktif aracını getir"""
    vehicle = await db.vehicles.find_one(
        {"user_id": current_user["id"], "is_active": True}, 
        {"_id": 0}
    )
    return vehicle

@api_router.get("/vehicles/{vehicle_id}")
async def get_vehicle(vehicle_id: str, current_user: dict = Depends(require_auth)):
    """Araç detayını getir"""
    vehicle = await db.vehicles.find_one(
        {"id": vehicle_id, "user_id": current_user["id"]}, 
        {"_id": 0}
    )
    if not vehicle:
        raise HTTPException(status_code=404, detail="Araç bulunamadı")
    return vehicle

@api_router.post("/vehicles")
async def create_vehicle(input: VehicleCreate, current_user: dict = Depends(require_auth)):
    """Yeni araç ekle"""
    # Eğer bu araç aktif olacaksa, diğerlerini pasif yap
    if input.is_active:
        await db.vehicles.update_many(
            {"user_id": current_user["id"]},
            {"$set": {"is_active": False}}
        )
    
    vehicle = Vehicle(
        user_id=current_user["id"],
        name=input.name,
        plate=input.plate,
        fuel_type=input.fuel_type,
        starting_km=input.starting_km,
        is_active=input.is_active
    )
    
    await db.vehicles.insert_one(vehicle.model_dump())
    return vehicle.model_dump()

@api_router.put("/vehicles/{vehicle_id}")
async def update_vehicle(vehicle_id: str, input: VehicleUpdate, current_user: dict = Depends(require_auth)):
    """Araç güncelle"""
    vehicle = await db.vehicles.find_one(
        {"id": vehicle_id, "user_id": current_user["id"]}, 
        {"_id": 0}
    )
    if not vehicle:
        raise HTTPException(status_code=404, detail="Araç bulunamadı")
    
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    
    # Eğer bu araç aktif yapılıyorsa, diğerlerini pasif yap
    if update_data.get("is_active") == True:
        await db.vehicles.update_many(
            {"user_id": current_user["id"], "id": {"$ne": vehicle_id}},
            {"$set": {"is_active": False}}
        )
    
    if update_data:
        await db.vehicles.update_one({"id": vehicle_id}, {"$set": update_data})
    
    updated = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    return updated

@api_router.delete("/vehicles/{vehicle_id}")
async def delete_vehicle(vehicle_id: str, current_user: dict = Depends(require_auth)):
    """Araç sil"""
    result = await db.vehicles.delete_one(
        {"id": vehicle_id, "user_id": current_user["id"]}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Araç bulunamadı")
    return {"message": "Araç silindi"}

# ===== YAKIT KAYITLARI =====
@api_router.get("/fuel-records")
async def get_fuel_records(
    vehicle_id: Optional[str] = None,
    limit: int = 50,
    current_user: dict = Depends(require_auth)
):
    """Yakıt kayıtlarını listele"""
    query = {"user_id": current_user["id"]}
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    
    records = await db.fuel_records.find(
        query, 
        {"_id": 0}
    ).sort("date", -1).to_list(limit)
    
    return records

@api_router.post("/fuel-records")
async def create_fuel_record(input: FuelRecordCreate, current_user: dict = Depends(require_auth)):
    """Yeni yakıt kaydı ekle ve otomatik hesaplamaları yap"""
    # Araç kontrolü
    vehicle = await db.vehicles.find_one(
        {"id": input.vehicle_id, "user_id": current_user["id"]}, 
        {"_id": 0}
    )
    if not vehicle:
        raise HTTPException(status_code=404, detail="Araç bulunamadı")
    
    # Bir önceki yakıt kaydını bul (aynı araç için)
    last_record = await db.fuel_records.find_one(
        {"vehicle_id": input.vehicle_id, "user_id": current_user["id"]},
        {"_id": 0},
        sort=[("current_km", -1)]
    )
    
    # Hesaplamalar
    distance_since_last = None
    consumption_per_100km = None
    cost_per_km = None
    
    if last_record and input.current_km > last_record.get("current_km", 0):
        distance_since_last = input.current_km - last_record["current_km"]
        if distance_since_last > 0 and input.liters > 0:
            consumption_per_100km = round((input.liters / distance_since_last) * 100, 2)
            cost_per_km = round(input.amount / distance_since_last, 3)
    
    record = FuelRecord(
        user_id=current_user["id"],
        vehicle_id=input.vehicle_id,
        date=input.date,
        current_km=input.current_km,
        liters=input.liters,
        amount=input.amount,
        note=input.note,
        distance_since_last=distance_since_last,
        consumption_per_100km=consumption_per_100km,
        cost_per_km=cost_per_km
    )
    
    await db.fuel_records.insert_one(record.model_dump())
    return record.model_dump()

@api_router.delete("/fuel-records/{record_id}")
async def delete_fuel_record(record_id: str, current_user: dict = Depends(require_auth)):
    """Yakıt kaydını sil"""
    result = await db.fuel_records.delete_one(
        {"id": record_id, "user_id": current_user["id"]}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    return {"message": "Yakıt kaydı silindi"}

# ===== GÜNLÜK KM TAKİBİ =====
@api_router.get("/daily-km")
async def get_daily_km_records(
    vehicle_id: Optional[str] = None,
    date: Optional[str] = None,
    limit: int = 30,
    current_user: dict = Depends(require_auth)
):
    """Günlük KM kayıtlarını listele"""
    query = {"user_id": current_user["id"]}
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    if date:
        query["date"] = date
    
    records = await db.daily_km_records.find(
        query, 
        {"_id": 0}
    ).sort("date", -1).to_list(limit)
    
    return records

@api_router.get("/daily-km/today")
async def get_today_km(current_user: dict = Depends(require_auth)):
    """Bugünkü KM kaydını getir (aktif araç için)"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Aktif aracı bul
    vehicle = await db.vehicles.find_one(
        {"user_id": current_user["id"], "is_active": True}, 
        {"_id": 0}
    )
    
    if not vehicle:
        return None
    
    record = await db.daily_km_records.find_one(
        {"user_id": current_user["id"], "vehicle_id": vehicle["id"], "date": today}, 
        {"_id": 0}
    )
    
    if record:
        record["vehicle"] = vehicle
    
    return record

async def calculate_avg_cost_per_km(user_id: str, vehicle_id: str) -> Optional[float]:
    """Son 30 günlük yakıt kayıtlarından ortalama km maliyetini hesapla"""
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    
    records = await db.fuel_records.find(
        {
            "user_id": user_id,
            "vehicle_id": vehicle_id,
            "date": {"$gte": thirty_days_ago},
            "cost_per_km": {"$ne": None}
        },
        {"_id": 0}
    ).to_list(100)
    
    if not records:
        return None
    
    costs = [r["cost_per_km"] for r in records if r.get("cost_per_km")]
    if not costs:
        return None
    
    return round(sum(costs) / len(costs), 3)

@api_router.post("/daily-km")
async def create_or_update_daily_km(
    input: DailyKmRecordCreate, 
    current_user: dict = Depends(require_auth)
):
    """Günlük KM kaydı oluştur veya güncelle"""
    # Araç kontrolü
    vehicle = await db.vehicles.find_one(
        {"id": input.vehicle_id, "user_id": current_user["id"]}, 
        {"_id": 0}
    )
    if not vehicle:
        raise HTTPException(status_code=404, detail="Araç bulunamadı")
    
    # Mevcut kayıt var mı?
    existing = await db.daily_km_records.find_one(
        {"user_id": current_user["id"], "vehicle_id": input.vehicle_id, "date": input.date},
        {"_id": 0}
    )
    
    # Hesaplamalar
    daily_km = None
    daily_cost = None
    avg_cost = await calculate_avg_cost_per_km(current_user["id"], input.vehicle_id)
    
    if input.end_km and input.start_km:
        daily_km = input.end_km - input.start_km
        if avg_cost and daily_km > 0:
            daily_cost = round(daily_km * avg_cost, 2)
    
    if existing:
        # Güncelle
        update_data = {
            "start_km": input.start_km,
            "daily_km": daily_km,
            "avg_cost_per_km": avg_cost,
            "daily_cost": daily_cost
        }
        if input.end_km:
            update_data["end_km"] = input.end_km
        
        await db.daily_km_records.update_one(
            {"id": existing["id"]},
            {"$set": update_data}
        )
        updated = await db.daily_km_records.find_one({"id": existing["id"]}, {"_id": 0})
        return updated
    else:
        # Yeni kayıt
        record = DailyKmRecord(
            user_id=current_user["id"],
            vehicle_id=input.vehicle_id,
            date=input.date,
            start_km=input.start_km,
            end_km=input.end_km,
            daily_km=daily_km,
            avg_cost_per_km=avg_cost,
            daily_cost=daily_cost
        )
        await db.daily_km_records.insert_one(record.model_dump())
        return record.model_dump()

@api_router.put("/daily-km/{record_id}")
async def update_daily_km(
    record_id: str,
    input: DailyKmRecordUpdate,
    current_user: dict = Depends(require_auth)
):
    """Günlük KM kaydını güncelle"""
    record = await db.daily_km_records.find_one(
        {"id": record_id, "user_id": current_user["id"]},
        {"_id": 0}
    )
    if not record:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    
    update_data = {}
    start_km = input.start_km if input.start_km is not None else record.get("start_km")
    end_km = input.end_km if input.end_km is not None else record.get("end_km")
    
    if input.start_km is not None:
        update_data["start_km"] = input.start_km
    if input.end_km is not None:
        update_data["end_km"] = input.end_km
    
    # Hesaplamalar
    if start_km and end_km:
        daily_km = end_km - start_km
        update_data["daily_km"] = daily_km
        
        avg_cost = await calculate_avg_cost_per_km(current_user["id"], record["vehicle_id"])
        update_data["avg_cost_per_km"] = avg_cost
        
        if avg_cost and daily_km > 0:
            update_data["daily_cost"] = round(daily_km * avg_cost, 2)
    
    if update_data:
        await db.daily_km_records.update_one({"id": record_id}, {"$set": update_data})
    
    updated = await db.daily_km_records.find_one({"id": record_id}, {"_id": 0})
    return updated

# ===== ARAÇ İSTATİSTİKLERİ =====
@api_router.get("/vehicle-stats/{vehicle_id}")
async def get_vehicle_stats(vehicle_id: str, current_user: dict = Depends(require_auth)):
    """Araç istatistiklerini getir"""
    vehicle = await db.vehicles.find_one(
        {"id": vehicle_id, "user_id": current_user["id"]},
        {"_id": 0}
    )
    if not vehicle:
        raise HTTPException(status_code=404, detail="Araç bulunamadı")
    
    # Tüm yakıt kayıtları
    fuel_records = await db.fuel_records.find(
        {"vehicle_id": vehicle_id, "user_id": current_user["id"]},
        {"_id": 0}
    ).sort("date", -1).to_list(1000)
    
    # Son 30 gün yakıt kayıtları
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    recent_records = [r for r in fuel_records if r.get("date", "") >= thirty_days_ago]
    
    # Hesaplamalar
    total_fuel_cost = sum(r.get("amount", 0) for r in fuel_records)
    total_liters = sum(r.get("liters", 0) for r in fuel_records)
    
    # Son 30 gün ortalamaları
    recent_costs_per_km = [r.get("cost_per_km") for r in recent_records if r.get("cost_per_km")]
    recent_consumption = [r.get("consumption_per_100km") for r in recent_records if r.get("consumption_per_100km")]
    
    avg_cost_per_km = round(sum(recent_costs_per_km) / len(recent_costs_per_km), 3) if recent_costs_per_km else None
    avg_consumption = round(sum(recent_consumption) / len(recent_consumption), 2) if recent_consumption else None
    
    # Bu ay yakıt gideri
    current_month = datetime.now(timezone.utc).strftime("%Y-%m")
    monthly_records = [r for r in fuel_records if r.get("date", "").startswith(current_month)]
    monthly_fuel_cost = sum(r.get("amount", 0) for r in monthly_records)
    
    return {
        "vehicle": vehicle,
        "total_fuel_cost": total_fuel_cost,
        "total_liters": total_liters,
        "monthly_fuel_cost": monthly_fuel_cost,
        "avg_cost_per_km": avg_cost_per_km,
        "avg_consumption_per_100km": avg_consumption,
        "fuel_record_count": len(fuel_records)
    }

# PDF Report endpoint - Profesyonel, Kompakt, Tablo Bazlı Format
@api_router.get("/report/pdf/{day_name}/{date}")
async def generate_daily_report_pdf(
    day_name: str, 
    date: str,
    current_user: dict = Depends(require_auth)
):
    """Generate professional daily visit report as PDF - compact table format"""
    
    # Kullanıcı bilgisi al
    user_name = current_user.get("name", "Satış Temsilcisi")
    user_email = current_user.get("email", "")
    
    # Get customers for this day (only user's customers)
    customers = await db.customers.find(
        {"visit_days": day_name, "user_id": current_user["id"]}, 
        {"_id": 0}
    ).to_list(1000)
    
    # Get visits for this date (only user's visits)
    visits = await db.visits.find({"date": date, "user_id": current_user["id"]}, {"_id": 0}).to_list(1000)
    
    # Apply migration to visits for status field
    for v in visits:
        migrate_visit_status(v)
    
    visits_map = {v["customer_id"]: v for v in visits}
    
    # Kategorize customers by visit status
    visited_customers = []
    not_visited_customers = []
    pending_customers = []
    
    for c in customers:
        visit = visits_map.get(c["id"], {})
        status = visit.get("status", "pending")
        if status == "visited":
            visited_customers.append((c, visit))
        elif status == "not_visited":
            not_visited_customers.append((c, visit))
        else:
            pending_customers.append((c, visit))
    
    # Get daily note (only user's note)
    daily_note = await db.daily_notes.find_one({"date": date, "user_id": current_user["id"]}, {"_id": 0})
    daily_note_text = daily_note.get("note", "") if daily_note else ""
    
    # Calculate stats
    total_count = len(customers)
    visited_count = len(visited_customers)
    not_visited_count = len(not_visited_customers)
    pending_count_stat = len(pending_customers)
    visit_rate = round((visited_count / total_count * 100), 1) if total_count > 0 else 0
    
    # Calculate payment stats
    total_payment = 0
    payment_count = 0
    for c, visit in visited_customers:
        if visit.get("payment_collected"):
            payment_count += 1
            total_payment += visit.get("payment_amount", 0) or 0
    
    # Get vehicle/km data
    daily_km_record = await db.daily_km_records.find_one(
        {"user_id": current_user["id"], "date": date},
        {"_id": 0}
    )
    vehicle = None
    if daily_km_record:
        vehicle = await db.vehicles.find_one(
            {"id": daily_km_record.get("vehicle_id")},
            {"_id": 0}
        )
    
    # Create PDF
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # Add Unicode font for Turkish characters
    pdf.add_font("DejaVu", "", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", uni=True)
    pdf.add_font("DejaVu", "B", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", uni=True)
    
    # =========================================================================
    # SAYFA 1: YÖNETİCİ ÖZETİ
    # =========================================================================
    pdf.add_page()
    
    # Header
    pdf.set_font("DejaVu", "B", 18)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(0, 10, "GÜNLÜK ZİYARET RAPORU", ln=True, align="C")
    
    pdf.set_font("DejaVu", "", 12)
    pdf.set_text_color(71, 85, 105)
    pdf.cell(0, 7, f"{day_name}, {date}", ln=True, align="C")
    
    pdf.ln(3)
    pdf.set_font("DejaVu", "B", 11)
    pdf.set_text_color(0, 85, 255)
    pdf.cell(0, 6, f"Satış Temsilcisi: {user_name}", ln=True, align="C")
    if user_email:
        pdf.set_font("DejaVu", "", 9)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(0, 5, user_email, ln=True, align="C")
    
    pdf.ln(8)
    
    # Özet Kutusu
    pdf.set_fill_color(248, 250, 252)
    pdf.set_draw_color(226, 232, 240)
    box_y = pdf.get_y()
    pdf.rect(10, box_y, 190, 55, "DF")
    
    # Sol Kolon - Ziyaret Özeti
    pdf.set_xy(15, box_y + 5)
    pdf.set_font("DejaVu", "B", 11)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(80, 6, "ZİYARET ÖZETİ")
    
    pdf.set_font("DejaVu", "", 10)
    pdf.set_xy(15, box_y + 14)
    pdf.set_text_color(71, 85, 105)
    pdf.cell(50, 5, "Planlanan Ziyaret:")
    pdf.set_font("DejaVu", "B", 10)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(30, 5, str(total_count))
    
    pdf.set_font("DejaVu", "", 10)
    pdf.set_xy(15, box_y + 22)
    pdf.set_text_color(22, 163, 74)
    pdf.cell(50, 5, "Ziyaret Edilen:")
    pdf.set_font("DejaVu", "B", 10)
    pdf.cell(30, 5, str(visited_count))
    
    pdf.set_font("DejaVu", "", 10)
    pdf.set_xy(15, box_y + 30)
    pdf.set_text_color(220, 38, 38)
    pdf.cell(50, 5, "Ziyaret Edilmeyen:")
    pdf.set_font("DejaVu", "B", 10)
    pdf.cell(30, 5, str(not_visited_count))
    
    pdf.set_font("DejaVu", "", 10)
    pdf.set_xy(15, box_y + 38)
    pdf.set_text_color(100, 116, 139)
    pdf.cell(50, 5, "Bekleyen:")
    pdf.set_font("DejaVu", "B", 10)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(30, 5, str(pending_count_stat))
    
    pdf.set_font("DejaVu", "B", 11)
    pdf.set_xy(15, box_y + 46)
    pdf.set_text_color(0, 85, 255)
    pdf.cell(50, 5, "Ziyaret Oranı:")
    pdf.cell(30, 5, f"%{visit_rate}")
    
    # Sağ Kolon - Tahsilat Özeti
    pdf.set_xy(110, box_y + 5)
    pdf.set_font("DejaVu", "B", 11)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(80, 6, "TAHSİLAT ÖZETİ")
    
    pdf.set_font("DejaVu", "", 10)
    pdf.set_xy(110, box_y + 14)
    pdf.set_text_color(71, 85, 105)
    pdf.cell(50, 5, "Tahsilat Yapılan:")
    pdf.set_font("DejaVu", "B", 10)
    pdf.set_text_color(22, 163, 74)
    pdf.cell(30, 5, f"{payment_count} müşteri")
    
    pdf.set_font("DejaVu", "B", 12)
    pdf.set_xy(110, box_y + 26)
    pdf.set_text_color(0, 85, 255)
    pdf.cell(50, 6, "Toplam Tahsilat:")
    pdf.cell(40, 6, f"{total_payment:,.2f} TL")
    
    # Araç bilgisi (varsa)
    if daily_km_record and vehicle:
        pdf.set_font("DejaVu", "", 9)
        pdf.set_xy(110, box_y + 38)
        pdf.set_text_color(71, 85, 105)
        daily_km = daily_km_record.get("daily_km")
        daily_cost = daily_km_record.get("daily_cost")
        pdf.cell(80, 5, f"Araç: {vehicle.get('name', '-')}")
        if daily_km:
            pdf.set_xy(110, box_y + 44)
            pdf.cell(40, 5, f"Günlük: {daily_km:,.0f} km")
            if daily_cost:
                pdf.cell(40, 5, f"Maliyet: {daily_cost:,.2f} TL")
    
    pdf.set_y(box_y + 60)
    
    # Gün Sonu Notu (varsa)
    if daily_note_text:
        pdf.ln(5)
        pdf.set_font("DejaVu", "B", 11)
        pdf.set_text_color(113, 63, 18)
        pdf.cell(0, 6, "GÜN SONU NOTU:", ln=True)
        
        pdf.set_fill_color(254, 252, 232)
        pdf.set_draw_color(250, 204, 21)
        note_y = pdf.get_y()
        pdf.rect(10, note_y, 190, 20, "DF")
        
        pdf.set_font("DejaVu", "", 9)
        pdf.set_text_color(113, 63, 18)
        pdf.set_xy(15, note_y + 3)
        pdf.multi_cell(180, 5, daily_note_text[:200])
    
    # =========================================================================
    # SAYFA 2: ZİYARET EDİLENLER TABLOSU
    # =========================================================================
    if visited_customers:
        pdf.add_page()
        
        pdf.set_font("DejaVu", "B", 14)
        pdf.set_text_color(22, 163, 74)
        pdf.cell(0, 10, f"ZİYARET EDİLEN MÜŞTERİLER ({len(visited_customers)})", ln=True)
        pdf.ln(2)
        
        # Tablo başlığı
        pdf.set_fill_color(220, 252, 231)
        pdf.set_draw_color(134, 239, 172)
        pdf.set_font("DejaVu", "B", 9)
        pdf.set_text_color(22, 101, 52)
        
        pdf.cell(8, 8, "#", border=1, fill=True, align="C")
        pdf.cell(52, 8, "Müşteri Adı", border=1, fill=True)
        pdf.cell(28, 8, "Bölge", border=1, fill=True)
        pdf.cell(35, 8, "Tahsilat", border=1, fill=True)
        pdf.cell(67, 8, "Not / Talep", border=1, fill=True, ln=True)
        
        # Tablo satırları
        pdf.set_font("DejaVu", "", 8)
        for i, (customer, visit) in enumerate(visited_customers, 1):
            # Sayfa kontrolü
            if pdf.get_y() > 260:
                pdf.add_page()
                # Başlık tekrar
                pdf.set_fill_color(220, 252, 231)
                pdf.set_font("DejaVu", "B", 9)
                pdf.set_text_color(22, 101, 52)
                pdf.cell(8, 8, "#", border=1, fill=True, align="C")
                pdf.cell(52, 8, "Müşteri Adı", border=1, fill=True)
                pdf.cell(28, 8, "Bölge", border=1, fill=True)
                pdf.cell(35, 8, "Tahsilat", border=1, fill=True)
                pdf.cell(67, 8, "Not / Talep", border=1, fill=True, ln=True)
                pdf.set_font("DejaVu", "", 8)
            
            pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(15, 23, 42)
            
            pdf.cell(8, 7, str(i), border=1, align="C")
            pdf.cell(52, 7, customer['name'][:28], border=1)
            pdf.cell(28, 7, customer['region'][:15], border=1)
            
            # Tahsilat
            if visit.get("payment_collected"):
                amount = visit.get("payment_amount", 0) or 0
                pdf.set_text_color(22, 163, 74)
                pdf.cell(35, 7, f"{amount:,.0f} TL", border=1)
            else:
                pdf.set_text_color(234, 88, 12)
                reason = visit.get("payment_skip_reason", "Yapılmadı")[:18]
                pdf.cell(35, 7, reason, border=1)
            
            # Not
            pdf.set_text_color(71, 85, 105)
            note_text = visit.get("customer_request") or visit.get("note") or "-"
            pdf.cell(67, 7, note_text[:35], border=1, ln=True)
    
    # =========================================================================
    # SAYFA 3: ZİYARET EDİLMEYENLER TABLOSU
    # =========================================================================
    if not_visited_customers:
        pdf.add_page()
        
        pdf.set_font("DejaVu", "B", 14)
        pdf.set_text_color(220, 38, 38)
        pdf.cell(0, 10, f"ZİYARET EDİLMEYEN MÜŞTERİLER ({len(not_visited_customers)})", ln=True)
        pdf.ln(2)
        
        # Tablo başlığı
        pdf.set_fill_color(254, 226, 226)
        pdf.set_draw_color(252, 165, 165)
        pdf.set_font("DejaVu", "B", 9)
        pdf.set_text_color(153, 27, 27)
        
        pdf.cell(8, 8, "#", border=1, fill=True, align="C")
        pdf.cell(60, 8, "Müşteri Adı", border=1, fill=True)
        pdf.cell(35, 8, "Bölge", border=1, fill=True)
        pdf.cell(87, 8, "Ziyaret Edilmeme Sebebi", border=1, fill=True, ln=True)
        
        # Tablo satırları
        pdf.set_font("DejaVu", "", 8)
        for i, (customer, visit) in enumerate(not_visited_customers, 1):
            if pdf.get_y() > 260:
                pdf.add_page()
                pdf.set_fill_color(254, 226, 226)
                pdf.set_font("DejaVu", "B", 9)
                pdf.set_text_color(153, 27, 27)
                pdf.cell(8, 8, "#", border=1, fill=True, align="C")
                pdf.cell(60, 8, "Müşteri Adı", border=1, fill=True)
                pdf.cell(35, 8, "Bölge", border=1, fill=True)
                pdf.cell(87, 8, "Ziyaret Edilmeme Sebebi", border=1, fill=True, ln=True)
                pdf.set_font("DejaVu", "", 8)
            
            pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(15, 23, 42)
            
            pdf.cell(8, 7, str(i), border=1, align="C")
            pdf.cell(60, 7, customer['name'][:32], border=1)
            pdf.cell(35, 7, customer['region'][:18], border=1)
            
            pdf.set_text_color(220, 38, 38)
            reason = visit.get("visit_skip_reason", "Belirtilmemiş")[:45]
            pdf.cell(87, 7, reason, border=1, ln=True)
    
    # =========================================================================
    # BEKLEYENLER (varsa kısa liste)
    # =========================================================================
    if pending_customers:
        if pdf.get_y() > 200:
            pdf.add_page()
        else:
            pdf.ln(10)
        
        pdf.set_font("DejaVu", "B", 12)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(0, 8, f"BEKLEYEN MÜŞTERİLER ({len(pending_customers)})", ln=True)
        
        pdf.set_font("DejaVu", "", 8)
        pdf.set_text_color(71, 85, 105)
        
        # Sadece isimlerini listele (kompakt)
        pending_names = [c['name'][:25] for c, v in pending_customers]
        pdf.multi_cell(0, 5, " • ".join(pending_names))
    
    # =========================================================================
    # FOOTER
    # =========================================================================
    pdf.ln(10)
    pdf.set_font("DejaVu", "", 7)
    pdf.set_text_color(148, 163, 184)
    report_date = datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')
    pdf.cell(0, 4, f"Rapor: {report_date} UTC | {user_name} | Satış Takip Sistemi", ln=True, align="C")
    
    # Output PDF
    pdf_output = io.BytesIO()
    pdf_content = pdf.output()
    pdf_output.write(pdf_content)
    pdf_output.seek(0)
    
    filename = f"ziyaret_raporu_{date}.pdf"
    
    return StreamingResponse(
        pdf_output,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

# =========================================================================
# DÖNEM RAPORU - Haftalık/Aylık Özet PDF
# =========================================================================
@api_router.get("/report/pdf/period/{period_type}")
async def generate_period_report_pdf(
    period_type: str,  # "weekly" or "monthly"
    start_date: str = None,  # Optional custom start date
    end_date: str = None,    # Optional custom end date
    current_user: dict = Depends(require_auth)
):
    """Generate weekly or monthly performance summary PDF report"""
    
    user_name = current_user.get("name", "Satış Temsilcisi")
    user_email = current_user.get("email", "")
    
    # Calculate date range
    today = datetime.now(timezone.utc).date()
    
    if start_date and end_date:
        period_start = datetime.strptime(start_date, "%Y-%m-%d").date()
        period_end = datetime.strptime(end_date, "%Y-%m-%d").date()
    elif period_type == "weekly":
        # Current week (Monday to Sunday)
        days_since_monday = today.weekday()
        period_start = today - timedelta(days=days_since_monday)
        period_end = period_start + timedelta(days=6)
    else:  # monthly
        period_start = today.replace(day=1)
        if today.month == 12:
            next_month = today.replace(year=today.year + 1, month=1, day=1)
        else:
            next_month = today.replace(month=today.month + 1, day=1)
        period_end = next_month - timedelta(days=1)
    
    start_str = period_start.isoformat()
    end_str = period_end.isoformat()
    
    # Get all visits in date range
    visits = await db.visits.find({
        "user_id": current_user["id"],
        "date": {"$gte": start_str, "$lte": end_str}
    }, {"_id": 0}).to_list(10000)
    
    # Apply migration for status
    for v in visits:
        migrate_visit_status(v)
    
    # Get all customers
    customers = await db.customers.find({"user_id": current_user["id"]}, {"_id": 0}).to_list(1000)
    
    # Get daily KM records
    daily_km_records = await db.daily_km_records.find({
        "user_id": current_user["id"],
        "date": {"$gte": start_str, "$lte": end_str}
    }, {"_id": 0}).to_list(1000)
    
    # Get fuel records
    fuel_records = await db.fuel_records.find({
        "user_id": current_user["id"],
        "date": {"$gte": start_str, "$lte": end_str}
    }, {"_id": 0}).to_list(1000)
    
    # Calculate statistics
    total_visits = len(visits)
    visited_count = sum(1 for v in visits if v.get("status") == "visited")
    not_visited_count = sum(1 for v in visits if v.get("status") == "not_visited")
    pending_count = total_visits - visited_count - not_visited_count
    visit_rate = round((visited_count / total_visits * 100), 1) if total_visits > 0 else 0
    
    # Payment stats
    total_payment = 0
    payment_count = 0
    payment_by_type = {"Nakit": 0, "Kredi Kartı": 0, "Havale/EFT": 0, "Çek": 0, "Diğer": 0}
    
    for v in visits:
        if v.get("payment_collected"):
            payment_count += 1
            amount = v.get("payment_amount", 0) or 0
            total_payment += amount
            ptype = v.get("payment_type", "Diğer")
            if ptype in payment_by_type:
                payment_by_type[ptype] += amount
            else:
                payment_by_type["Diğer"] += amount
    
    # Working days (unique dates with visits)
    unique_dates = set(v.get("date") for v in visits if v.get("date"))
    working_days = len(unique_dates)
    
    # Daily averages
    avg_daily_visits = round(visited_count / working_days, 1) if working_days > 0 else 0
    avg_daily_payment = round(total_payment / working_days, 2) if working_days > 0 else 0
    
    # Vehicle/Fuel stats
    total_km = sum(r.get("daily_km", 0) or 0 for r in daily_km_records)
    total_fuel_cost = sum(r.get("amount", 0) or 0 for r in fuel_records)
    avg_km_cost = round(total_fuel_cost / total_km, 3) if total_km > 0 else 0
    
    # Daily data for charts
    daily_data = {}
    for v in visits:
        date = v.get("date")
        if date:
            if date not in daily_data:
                daily_data[date] = {"visited": 0, "not_visited": 0, "payment": 0}
            if v.get("status") == "visited":
                daily_data[date]["visited"] += 1
            elif v.get("status") == "not_visited":
                daily_data[date]["not_visited"] += 1
            if v.get("payment_collected"):
                daily_data[date]["payment"] += v.get("payment_amount", 0) or 0
    
    # Sort daily data by date
    sorted_dates = sorted(daily_data.keys())
    
    # Create PDF
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_font("DejaVu", "", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", uni=True)
    pdf.add_font("DejaVu", "B", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", uni=True)
    
    # =========================================================================
    # SAYFA 1: DÖNEM ÖZETİ
    # =========================================================================
    pdf.add_page()
    
    # Header
    period_label = "HAFTALIK" if period_type == "weekly" else "AYLIK"
    pdf.set_font("DejaVu", "B", 18)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(0, 10, f"{period_label} PERFORMANS RAPORU", ln=True, align="C")
    
    pdf.set_font("DejaVu", "", 11)
    pdf.set_text_color(71, 85, 105)
    pdf.cell(0, 7, f"{period_start.strftime('%d.%m.%Y')} - {period_end.strftime('%d.%m.%Y')}", ln=True, align="C")
    
    pdf.ln(3)
    pdf.set_font("DejaVu", "B", 11)
    pdf.set_text_color(0, 85, 255)
    pdf.cell(0, 6, f"Satış Temsilcisi: {user_name}", ln=True, align="C")
    if user_email:
        pdf.set_font("DejaVu", "", 9)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(0, 5, user_email, ln=True, align="C")
    
    pdf.ln(8)
    
    # Çalışma Özeti
    pdf.set_fill_color(248, 250, 252)
    pdf.set_draw_color(226, 232, 240)
    box_y = pdf.get_y()
    pdf.rect(10, box_y, 190, 25, "DF")
    
    pdf.set_font("DejaVu", "B", 10)
    pdf.set_text_color(15, 23, 42)
    pdf.set_xy(15, box_y + 5)
    pdf.cell(60, 6, f"Çalışılan Gün: {working_days}")
    pdf.cell(60, 6, f"Toplam Müşteri: {len(customers)}")
    pdf.cell(60, 6, f"Toplam Ziyaret Kaydı: {total_visits}")
    
    pdf.set_y(box_y + 30)
    
    # ===== ZİYARET PERFORMANSI =====
    pdf.set_font("DejaVu", "B", 12)
    pdf.set_text_color(22, 163, 74)
    pdf.cell(0, 8, "ZİYARET PERFORMANSI", ln=True)
    
    pdf.set_fill_color(220, 252, 231)
    pdf.set_draw_color(134, 239, 172)
    box_y = pdf.get_y()
    pdf.rect(10, box_y, 190, 40, "DF")
    
    pdf.set_font("DejaVu", "", 10)
    pdf.set_text_color(22, 101, 52)
    
    pdf.set_xy(15, box_y + 5)
    pdf.cell(60, 6, f"Ziyaret Edilen: {visited_count}")
    pdf.cell(60, 6, f"Ziyaret Edilmeyen: {not_visited_count}")
    pdf.cell(60, 6, f"Bekleyen: {pending_count}")
    
    pdf.set_xy(15, box_y + 14)
    pdf.set_font("DejaVu", "B", 11)
    pdf.cell(60, 6, f"Ziyaret Oranı: %{visit_rate}")
    pdf.set_font("DejaVu", "", 10)
    pdf.cell(60, 6, f"Günlük Ort. Ziyaret: {avg_daily_visits}")
    
    pdf.set_y(box_y + 45)
    
    # ===== TAHSİLAT PERFORMANSI =====
    pdf.set_font("DejaVu", "B", 12)
    pdf.set_text_color(0, 85, 255)
    pdf.cell(0, 8, "TAHSİLAT PERFORMANSI", ln=True)
    
    pdf.set_fill_color(219, 234, 254)
    pdf.set_draw_color(147, 197, 253)
    box_y = pdf.get_y()
    pdf.rect(10, box_y, 190, 50, "DF")
    
    pdf.set_font("DejaVu", "", 10)
    pdf.set_text_color(30, 64, 175)
    
    pdf.set_xy(15, box_y + 5)
    pdf.cell(90, 6, f"Tahsilat Yapılan Müşteri: {payment_count}")
    pdf.cell(90, 6, f"Günlük Ort. Tahsilat: {avg_daily_payment:,.2f} TL")
    
    pdf.set_xy(15, box_y + 14)
    pdf.set_font("DejaVu", "B", 14)
    pdf.cell(0, 8, f"TOPLAM TAHSİLAT: {total_payment:,.2f} TL", ln=True)
    
    # Payment breakdown
    pdf.set_font("DejaVu", "", 9)
    pdf.set_xy(15, box_y + 28)
    col_x = 15
    for ptype, amount in payment_by_type.items():
        if amount > 0:
            pdf.set_xy(col_x, box_y + 28)
            pdf.cell(45, 5, f"{ptype}: {amount:,.0f} TL")
            col_x += 45
            if col_x > 160:
                col_x = 15
                pdf.ln(6)
    
    pdf.set_y(box_y + 55)
    
    # ===== ARAÇ/YAKIT ÖZETİ =====
    if total_km > 0 or total_fuel_cost > 0:
        pdf.set_font("DejaVu", "B", 12)
        pdf.set_text_color(113, 63, 18)
        pdf.cell(0, 8, "ARAÇ / YAKIT ÖZETİ", ln=True)
        
        pdf.set_fill_color(254, 252, 232)
        pdf.set_draw_color(250, 204, 21)
        box_y = pdf.get_y()
        pdf.rect(10, box_y, 190, 25, "DF")
        
        pdf.set_font("DejaVu", "", 10)
        pdf.set_text_color(113, 63, 18)
        
        pdf.set_xy(15, box_y + 5)
        pdf.cell(60, 6, f"Toplam KM: {total_km:,.0f} km")
        pdf.cell(60, 6, f"Toplam Yakıt: {total_fuel_cost:,.2f} TL")
        pdf.cell(60, 6, f"Ort. KM Maliyeti: {avg_km_cost:.3f} TL/km")
        
        pdf.set_y(box_y + 30)
    
    # =========================================================================
    # SAYFA 2: GÜNLÜK DETAY TABLOSU
    # =========================================================================
    if sorted_dates:
        pdf.add_page()
        
        pdf.set_font("DejaVu", "B", 14)
        pdf.set_text_color(15, 23, 42)
        pdf.cell(0, 10, "GÜNLÜK PERFORMANS DETAYI", ln=True)
        pdf.ln(2)
        
        # Tablo başlığı
        pdf.set_fill_color(241, 245, 249)
        pdf.set_draw_color(203, 213, 225)
        pdf.set_font("DejaVu", "B", 9)
        pdf.set_text_color(15, 23, 42)
        
        pdf.cell(30, 8, "Tarih", border=1, fill=True, align="C")
        pdf.cell(25, 8, "Gün", border=1, fill=True, align="C")
        pdf.cell(30, 8, "Ziyaret", border=1, fill=True, align="C")
        pdf.cell(30, 8, "Edilmedi", border=1, fill=True, align="C")
        pdf.cell(35, 8, "Oran", border=1, fill=True, align="C")
        pdf.cell(40, 8, "Tahsilat", border=1, fill=True, align="C", ln=True)
        
        # Türkçe gün isimleri
        gun_isimleri = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
        
        pdf.set_font("DejaVu", "", 8)
        for date_str in sorted_dates:
            if pdf.get_y() > 260:
                pdf.add_page()
                pdf.set_fill_color(241, 245, 249)
                pdf.set_font("DejaVu", "B", 9)
                pdf.cell(30, 8, "Tarih", border=1, fill=True, align="C")
                pdf.cell(25, 8, "Gün", border=1, fill=True, align="C")
                pdf.cell(30, 8, "Ziyaret", border=1, fill=True, align="C")
                pdf.cell(30, 8, "Edilmedi", border=1, fill=True, align="C")
                pdf.cell(35, 8, "Oran", border=1, fill=True, align="C")
                pdf.cell(40, 8, "Tahsilat", border=1, fill=True, align="C", ln=True)
                pdf.set_font("DejaVu", "", 8)
            
            data = daily_data[date_str]
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            gun_adi = gun_isimleri[date_obj.weekday()]
            
            total_day = data["visited"] + data["not_visited"]
            day_rate = round((data["visited"] / total_day * 100), 0) if total_day > 0 else 0
            
            pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(15, 23, 42)
            
            pdf.cell(30, 7, date_obj.strftime("%d.%m"), border=1, align="C")
            pdf.cell(25, 7, gun_adi[:3], border=1, align="C")
            
            pdf.set_text_color(22, 163, 74)
            pdf.cell(30, 7, str(data["visited"]), border=1, align="C")
            
            pdf.set_text_color(220, 38, 38)
            pdf.cell(30, 7, str(data["not_visited"]), border=1, align="C")
            
            pdf.set_text_color(0, 85, 255)
            pdf.cell(35, 7, f"%{day_rate:.0f}", border=1, align="C")
            
            pdf.set_text_color(15, 23, 42)
            pdf.cell(40, 7, f"{data['payment']:,.0f} TL", border=1, align="C", ln=True)
        
        # Toplam satırı
        pdf.set_font("DejaVu", "B", 9)
        pdf.set_fill_color(241, 245, 249)
        pdf.cell(30, 8, "TOPLAM", border=1, fill=True, align="C")
        pdf.cell(25, 8, f"{working_days} gün", border=1, fill=True, align="C")
        pdf.set_text_color(22, 163, 74)
        pdf.cell(30, 8, str(visited_count), border=1, fill=True, align="C")
        pdf.set_text_color(220, 38, 38)
        pdf.cell(30, 8, str(not_visited_count), border=1, fill=True, align="C")
        pdf.set_text_color(0, 85, 255)
        pdf.cell(35, 8, f"%{visit_rate}", border=1, fill=True, align="C")
        pdf.set_text_color(15, 23, 42)
        pdf.cell(40, 8, f"{total_payment:,.0f} TL", border=1, fill=True, align="C", ln=True)
    
    # Footer
    pdf.ln(10)
    pdf.set_font("DejaVu", "", 7)
    pdf.set_text_color(148, 163, 184)
    report_date = datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')
    pdf.cell(0, 4, f"Rapor: {report_date} UTC | {user_name} | Satış Takip Sistemi", ln=True, align="C")
    
    # Output PDF
    pdf_output = io.BytesIO()
    pdf_content = pdf.output()
    pdf_output.write(pdf_content)
    pdf_output.seek(0)
    
    period_label_file = "haftalik" if period_type == "weekly" else "aylik"
    filename = f"performans_raporu_{period_label_file}_{period_start.strftime('%Y%m%d')}.pdf"
    
    return StreamingResponse(
        pdf_output,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
